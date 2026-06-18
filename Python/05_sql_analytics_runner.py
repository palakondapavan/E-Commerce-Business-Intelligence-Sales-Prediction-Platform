"""
============================================================
 E-Commerce BI & Sales Prediction Platform
 Script 05 — SQL Analytics Runner (SQLite-based)
 Runs all 25 business queries & exports results to Excel
============================================================
"""

import os, sqlite3, warnings
warnings.filterwarnings("ignore")

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime

BASE    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CLN_CSV = os.path.join(BASE, "Dataset", "Amazon_Sale_Report_Cleaned.csv")
DB_PATH = os.path.join(BASE, "Dataset", "ecommerce_bi.db")
RPT_DIR = os.path.join(BASE, "Reports")
SQL_DIR = os.path.join(BASE, "SQL")
XLS_OUT = os.path.join(RPT_DIR, "SQL_Analytics_Report.xlsx")
os.makedirs(RPT_DIR, exist_ok=True)

BG    = "#0f1117"
PANEL = "#1a1d2e"
PAL   = ["#00c8ff","#ff4d6d","#ffd166","#06d6a0","#c77dff",
         "#ff9f1c","#a8dadc","#f72585","#4cc9f0","#b5e48c"]

print("\n" + "="*65)
print("  SQL ANALYTICS RUNNER — E-Commerce BI Platform")
print("="*65)

# ═══════════════════════════════════════════════════════════
#  STEP 1 — Load CSV into SQLite (in-memory star schema)
# ═══════════════════════════════════════════════════════════
print("\n  [1/5] Building SQLite in-memory database …")

df = pd.read_csv(CLN_CSV, parse_dates=["date"], low_memory=False)

# Normalise columns expected by SQL queries
df["order_date"]   = df["date"].dt.date.astype(str)
df["year"]         = df["date"].dt.year
df["month"]        = df["date"].dt.month
df["month_name"]   = df["date"].dt.strftime("%b")
df["quarter"]      = df["date"].dt.quarter
df["day_name"]     = df["date"].dt.day_name()
df["week_number"]  = df["date"].dt.isocalendar().week.astype(int)
df["unit_price"]   = df["unit_price"].fillna(df["amount"])
df["promotion_flag"] = df["promotion_ids"].apply(
    lambda x: "With Promotion" if x != "No Promotion" else "No Promotion")

conn = sqlite3.connect(DB_PATH)
df.to_sql("orders", conn, if_exists="replace", index=False)

# Create handy views
conn.executescript("""
    DROP VIEW IF EXISTS v_delivered;
    CREATE VIEW v_delivered AS
    SELECT * FROM orders
    WHERE LOWER(status) IN ('delivered','shipped','shipped - delivered to buyer');

    DROP VIEW IF EXISTS v_monthly;
    CREATE VIEW v_monthly AS
    SELECT year, month, month_name,
           COUNT(DISTINCT order_id)  AS total_orders,
           SUM(qty)                  AS total_qty,
           ROUND(SUM(amount),2)      AS total_revenue,
           ROUND(AVG(amount),2)      AS avg_order_value
    FROM orders
    GROUP BY year, month, month_name;
""")
conn.commit()

total = pd.read_sql("SELECT COUNT(*) AS n FROM orders", conn).iloc[0,0]
print(f"  Loaded {total:,} rows into SQLite: {DB_PATH}")

# ═══════════════════════════════════════════════════════════
#  STEP 2 — Define & Run 25 Business Queries
# ═══════════════════════════════════════════════════════════
print("\n  [2/5] Running 25 business analysis queries …")

QUERIES = {

# ── A: OVERALL KPIs ────────────────────────────────────────────
"Q01_Overall_KPIs": (
    "Overall Business KPIs",
    """
    SELECT
        COUNT(DISTINCT order_id)            AS total_orders,
        SUM(qty)                            AS total_quantity_sold,
        ROUND(SUM(amount),2)                AS total_revenue_inr,
        ROUND(AVG(amount),2)                AS avg_order_value,
        ROUND(SUM(amount)/COUNT(DISTINCT order_id),2) AS revenue_per_order,
        ROUND(MIN(amount),2)                AS min_order_value,
        ROUND(MAX(amount),2)                AS max_order_value
    FROM orders
    """
),

"Q02_Order_Status_Analysis": (
    "Order Status Distribution",
    """
    SELECT
        status,
        COUNT(*)                                           AS order_count,
        ROUND(COUNT(*)*100.0/SUM(COUNT(*)) OVER(),2)      AS pct_of_total,
        ROUND(SUM(amount),2)                              AS total_revenue,
        ROUND(AVG(amount),2)                              AS avg_order_value
    FROM orders
    GROUP BY status
    ORDER BY order_count DESC
    """
),

"Q03_B2B_vs_B2C": (
    "B2B vs B2C Revenue Split",
    """
    SELECT
        b2b                                              AS customer_type,
        COUNT(DISTINCT order_id)                         AS orders,
        SUM(qty)                                         AS qty_sold,
        ROUND(SUM(amount),2)                             AS total_revenue,
        ROUND(AVG(amount),2)                             AS avg_order_value,
        ROUND(SUM(amount)*100.0/(SELECT SUM(amount) FROM orders),2) AS revenue_pct
    FROM orders
    GROUP BY b2b
    ORDER BY total_revenue DESC
    """
),

# ── B: TIME ANALYSIS ───────────────────────────────────────────
"Q04_Monthly_Revenue_Trend": (
    "Monthly Revenue Trend",
    """
    SELECT year, month, month_name,
           total_orders, total_qty,
           total_revenue, avg_order_value,
           ROUND(total_revenue - LAG(total_revenue,1,total_revenue)
                 OVER (ORDER BY year,month),2) AS mom_revenue_change,
           ROUND((total_revenue - LAG(total_revenue,1,total_revenue)
                 OVER (ORDER BY year,month))*100.0
                 / LAG(total_revenue,1,total_revenue) OVER (ORDER BY year,month),2)
                                               AS mom_growth_pct
    FROM v_monthly
    ORDER BY year, month
    """
),

"Q05_Quarterly_Performance": (
    "Quarterly Performance",
    """
    SELECT year, quarter,
           COUNT(DISTINCT order_id)  AS total_orders,
           SUM(qty)                  AS total_qty,
           ROUND(SUM(amount),2)      AS total_revenue,
           ROUND(AVG(amount),2)      AS avg_order_value
    FROM orders
    GROUP BY year, quarter
    ORDER BY year, quarter
    """
),

"Q06_Day_of_Week_Pattern": (
    "Day-of-Week Revenue Pattern",
    """
    SELECT day_name,
           COUNT(DISTINCT order_id)  AS total_orders,
           SUM(qty)                  AS total_qty,
           ROUND(SUM(amount),2)      AS total_revenue,
           ROUND(AVG(amount),2)      AS avg_order_value
    FROM orders
    GROUP BY day_name
    ORDER BY CASE day_name
        WHEN 'Monday' THEN 1 WHEN 'Tuesday' THEN 2 WHEN 'Wednesday' THEN 3
        WHEN 'Thursday' THEN 4 WHEN 'Friday' THEN 5
        WHEN 'Saturday' THEN 6 WHEN 'Sunday' THEN 7 END
    """
),

"Q07_Weekly_Revenue": (
    "Weekly Revenue Trend",
    """
    SELECT year, week_number,
           COUNT(DISTINCT order_id)  AS total_orders,
           ROUND(SUM(amount),2)      AS weekly_revenue,
           ROUND(AVG(amount),2)      AS avg_order_value
    FROM orders
    GROUP BY year, week_number
    ORDER BY year, week_number
    """
),

# ── C: PRODUCT & CATEGORY ──────────────────────────────────────
"Q08_Revenue_by_Category": (
    "Revenue by Category",
    """
    SELECT category,
           COUNT(DISTINCT order_id)                         AS total_orders,
           SUM(qty)                                         AS total_qty,
           ROUND(SUM(amount),2)                             AS total_revenue,
           ROUND(AVG(amount),2)                             AS avg_order_value,
           ROUND(SUM(amount)*100.0/(SELECT SUM(amount) FROM orders),2) AS revenue_pct
    FROM orders
    GROUP BY category
    ORDER BY total_revenue DESC
    """
),

"Q09_Top10_Products_Revenue": (
    "Top 10 Products by Revenue",
    """
    SELECT sku, style, category,
           COUNT(DISTINCT order_id)  AS total_orders,
           SUM(qty)                  AS qty_sold,
           ROUND(SUM(amount),2)      AS total_revenue,
           ROUND(AVG(unit_price),2)  AS avg_unit_price
    FROM orders
    GROUP BY sku, style, category
    ORDER BY total_revenue DESC
    LIMIT 10
    """
),

"Q10_Top10_Products_Qty": (
    "Top 10 Products by Quantity",
    """
    SELECT sku, style, category,
           SUM(qty)             AS qty_sold,
           ROUND(SUM(amount),2) AS total_revenue,
           ROUND(AVG(amount),2) AS avg_order_value
    FROM orders
    GROUP BY sku, style, category
    ORDER BY qty_sold DESC
    LIMIT 10
    """
),

"Q11_Category_Monthly_Trend": (
    "Category Monthly Breakdown",
    """
    SELECT category, month_name, month,
           COUNT(DISTINCT order_id)  AS orders,
           ROUND(SUM(amount),2)      AS revenue,
           SUM(qty)                  AS qty_sold
    FROM orders
    GROUP BY category, month_name, month
    ORDER BY category, month
    """
),

"Q12_Size_Distribution": (
    "Size Distribution by Category",
    """
    SELECT category, size,
           COUNT(DISTINCT order_id)  AS orders,
           SUM(qty)                  AS qty_sold,
           ROUND(SUM(amount),2)      AS revenue
    FROM orders
    GROUP BY category, size
    ORDER BY category, qty_sold DESC
    """
),

# ── D: GEOGRAPHY ───────────────────────────────────────────────
"Q13_Top_States_Revenue": (
    "Top States by Revenue",
    """
    SELECT ship_state,
           COUNT(DISTINCT order_id)                         AS total_orders,
           SUM(qty)                                         AS total_qty,
           ROUND(SUM(amount),2)                             AS total_revenue,
           ROUND(AVG(amount),2)                             AS avg_order_value,
           ROUND(SUM(amount)*100.0/(SELECT SUM(amount) FROM orders),2) AS revenue_pct
    FROM orders
    WHERE ship_state != 'Unknown'
    GROUP BY ship_state
    ORDER BY total_revenue DESC
    LIMIT 15
    """
),

"Q14_Top_Cities_Orders": (
    "Top 15 Cities by Orders",
    """
    SELECT ship_city, ship_state,
           COUNT(DISTINCT order_id)  AS total_orders,
           SUM(qty)                  AS qty_sold,
           ROUND(SUM(amount),2)      AS total_revenue,
           ROUND(AVG(amount),2)      AS avg_order_value
    FROM orders
    WHERE ship_city != 'Unknown'
    GROUP BY ship_city, ship_state
    ORDER BY total_orders DESC
    LIMIT 15
    """
),

"Q15_State_Cancellation_Rate": (
    "State-Level Cancellation Rates",
    """
    SELECT ship_state,
           COUNT(*)                                               AS total_orders,
           SUM(CASE WHEN status='Delivered' THEN 1 ELSE 0 END)   AS delivered,
           SUM(CASE WHEN status='Shipped'   THEN 1 ELSE 0 END)   AS shipped,
           SUM(CASE WHEN status='Cancelled' THEN 1 ELSE 0 END)   AS cancelled,
           SUM(CASE WHEN status='Returned'  THEN 1 ELSE 0 END)   AS returned,
           ROUND(SUM(CASE WHEN status='Cancelled' THEN 1 ELSE 0 END)*100.0/COUNT(*),2)
                                                                  AS cancel_rate_pct
    FROM orders
    WHERE ship_state != 'Unknown'
    GROUP BY ship_state
    ORDER BY cancel_rate_pct DESC
    LIMIT 15
    """
),

# ── E: CHANNEL & FULFILMENT ────────────────────────────────────
"Q16_Sales_Channel_Revenue": (
    "Sales Channel Revenue",
    """
    SELECT sales_channel,
           COUNT(DISTINCT order_id)                         AS total_orders,
           SUM(qty)                                         AS qty_sold,
           ROUND(SUM(amount),2)                             AS total_revenue,
           ROUND(AVG(amount),2)                             AS avg_order_value,
           ROUND(SUM(amount)*100.0/(SELECT SUM(amount) FROM orders),2) AS revenue_pct
    FROM orders
    GROUP BY sales_channel
    ORDER BY total_revenue DESC
    """
),

"Q17_Fulfilment_Analysis": (
    "Fulfilment Method Analysis",
    """
    SELECT fulfilment,
           COUNT(DISTINCT order_id)  AS total_orders,
           ROUND(SUM(amount),2)      AS total_revenue,
           ROUND(AVG(amount),2)      AS avg_order_value,
           SUM(CASE WHEN status='Delivered' THEN 1 ELSE 0 END) AS delivered,
           SUM(CASE WHEN status='Cancelled' THEN 1 ELSE 0 END) AS cancelled,
           ROUND(SUM(CASE WHEN status='Cancelled' THEN 1 ELSE 0 END)*100.0
                 /COUNT(*),2) AS cancel_rate_pct
    FROM orders
    GROUP BY fulfilment
    """
),

"Q18_Service_Level_Analysis": (
    "Shipping Service Level Analysis",
    """
    SELECT ship_service_level,
           COUNT(DISTINCT order_id)  AS total_orders,
           ROUND(SUM(amount),2)      AS total_revenue,
           ROUND(AVG(amount),2)      AS avg_order_value,
           ROUND(AVG(qty),2)         AS avg_qty_per_order
    FROM orders
    GROUP BY ship_service_level
    ORDER BY total_revenue DESC
    """
),

"Q19_Promotion_Impact": (
    "Promotion Impact Analysis",
    """
    SELECT promotion_flag AS promo_segment,
           COUNT(DISTINCT order_id)  AS total_orders,
           ROUND(SUM(amount),2)      AS total_revenue,
           ROUND(AVG(amount),2)      AS avg_order_value,
           SUM(qty)                  AS total_qty
    FROM orders
    GROUP BY promotion_flag
    """
),

# ── F: CUSTOMER INSIGHTS ───────────────────────────────────────
"Q20_High_Value_Orders": (
    "High Value Order Segments",
    """
    WITH order_totals AS (
        SELECT order_id, SUM(amount) AS order_total
        FROM orders GROUP BY order_id
    ),
    quartiled AS (
        SELECT order_id, order_total,
               NTILE(4) OVER (ORDER BY order_total) AS quartile
        FROM order_totals
    )
    SELECT quartile AS value_tier,
           COUNT(*)                  AS orders,
           ROUND(MIN(order_total),2) AS min_value,
           ROUND(MAX(order_total),2) AS max_value,
           ROUND(AVG(order_total),2) AS avg_value,
           ROUND(SUM(order_total),2) AS total_revenue,
           ROUND(SUM(order_total)*100.0/(SELECT SUM(amount) FROM orders),2) AS revenue_pct
    FROM quartiled
    GROUP BY quartile
    ORDER BY quartile
    """
),

"Q21_Category_State_Matrix": (
    "Category × Top-5 State Revenue Matrix",
    """
    SELECT ship_state, category,
           ROUND(SUM(amount),2) AS revenue,
           COUNT(DISTINCT order_id) AS orders
    FROM orders
    WHERE ship_state IN (
        SELECT ship_state FROM orders
        WHERE ship_state != 'Unknown'
        GROUP BY ship_state
        ORDER BY SUM(amount) DESC LIMIT 5
    )
    GROUP BY ship_state, category
    ORDER BY ship_state, revenue DESC
    """
),

"Q22_Cumulative_Revenue": (
    "Cumulative Revenue (YTD)",
    """
    SELECT year, month, month_name, total_revenue,
           ROUND(SUM(total_revenue) OVER
                 (PARTITION BY year ORDER BY month),2) AS cumulative_ytd
    FROM v_monthly
    ORDER BY year, month
    """
),

"Q23_Return_Rate_by_Category": (
    "Return Rate by Category",
    """
    SELECT category,
           COUNT(*)                                           AS total_orders,
           SUM(CASE WHEN status='Returned' THEN 1 ELSE 0 END) AS returned_orders,
           ROUND(SUM(CASE WHEN status='Returned' THEN 1 ELSE 0 END)*100.0/COUNT(*),2)
                                                              AS return_rate_pct,
           ROUND(SUM(CASE WHEN status='Returned' THEN amount ELSE 0 END),2)
                                                              AS returned_revenue
    FROM orders
    GROUP BY category
    ORDER BY return_rate_pct DESC
    """
),

"Q24_Revenue_by_Courier": (
    "Revenue & Orders by Courier",
    """
    SELECT courier_status AS courier,
           COUNT(DISTINCT order_id)  AS total_orders,
           ROUND(SUM(amount),2)      AS total_revenue,
           ROUND(AVG(amount),2)      AS avg_order_value,
           SUM(qty)                  AS total_qty
    FROM orders
    WHERE courier_status != 'Unknown'
    GROUP BY courier_status
    ORDER BY total_revenue DESC
    """
),

"Q25_Executive_Summary": (
    "Executive Summary Dashboard",
    """
    SELECT 'Total Revenue (INR)'    AS metric,
           ROUND(SUM(amount),2)     AS value FROM orders
    UNION ALL
    SELECT 'Total Orders',           COUNT(DISTINCT order_id) FROM orders
    UNION ALL
    SELECT 'Total Qty Sold',         SUM(qty)                 FROM orders
    UNION ALL
    SELECT 'Avg Order Value (INR)',   ROUND(AVG(amount),2)     FROM orders
    UNION ALL
    SELECT 'Delivered Orders',
           SUM(CASE WHEN LOWER(status) LIKE '%deliver%' THEN 1 ELSE 0 END) FROM orders
    UNION ALL
    SELECT 'Cancelled Orders',
           SUM(CASE WHEN status='Cancelled' THEN 1 ELSE 0 END) FROM orders
    UNION ALL
    SELECT 'Cancellation Rate (%)',
           ROUND(SUM(CASE WHEN status='Cancelled' THEN 1 ELSE 0 END)*100.0/COUNT(*),2) FROM orders
    UNION ALL
    SELECT 'Return Rate (%)',
           ROUND(SUM(CASE WHEN status='Returned' THEN 1 ELSE 0 END)*100.0/COUNT(*),2)  FROM orders
    UNION ALL
    SELECT 'B2B Revenue Share (%)',
           ROUND(SUM(CASE WHEN b2b='Yes' THEN amount ELSE 0 END)*100.0/SUM(amount),2)  FROM orders
    UNION ALL
    SELECT 'Unique SKUs',            COUNT(DISTINCT sku)      FROM orders
    UNION ALL
    SELECT 'Active States',          COUNT(DISTINCT ship_state) FROM orders WHERE ship_state!='Unknown'
    UNION ALL
    SELECT 'Active Cities',          COUNT(DISTINCT ship_city)  FROM orders WHERE ship_city!='Unknown'
    """
),
}

results = {}
for key, (title, sql) in QUERIES.items():
    try:
        df_res = pd.read_sql(sql, conn)
        results[key] = (title, df_res)
        print(f"    ✅  {key} — {len(df_res)} rows")
    except Exception as e:
        print(f"    ❌  {key} — ERROR: {e}")

conn.close()
print(f"\n  Executed {len(results)}/25 queries successfully")

# ═══════════════════════════════════════════════════════════
#  STEP 3 — Export to Styled Excel Workbook
# ═══════════════════════════════════════════════════════════
print("\n  [3/5] Building styled Excel report …")

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ── Style helpers ─────────────────────────────────────────
def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))

def thin_border():
    s = Side(style="thin", color="1a1d2e")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL  = hex_fill("0d0d1a")
HDR_FONT  = Font(name="Calibri", bold=True, color="00c8ff", size=11)
CELL_FILL_ODD  = hex_fill("13132b")
CELL_FILL_EVEN = hex_fill("1c1c3a")
CELL_FONT      = Font(name="Calibri", color="e8e8f0", size=10)
TITLE_FONT     = Font(name="Calibri", bold=True, color="ffd166", size=14)
SUB_FONT       = Font(name="Calibri", color="aaaacc", size=10)

def write_query_sheet(wb, sheet_name, title, df_data):
    """Write a single query result to a styled sheet."""
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.sheet_properties.tabColor = "00c8ff"

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(max(len(df_data.columns),1))}1")
    ws["A1"] = title
    ws["A1"].font  = TITLE_FONT
    ws["A1"].fill  = HDR_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub header
    ws.merge_cells(f"A2:{get_column_letter(max(len(df_data.columns),1))}2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Rows: {len(df_data)}"
    ws["A2"].font  = SUB_FONT
    ws["A2"].fill  = hex_fill("0f1117")
    ws.row_dimensions[2].height = 18

    # Column headers (row 3)
    for c_idx, col in enumerate(df_data.columns, 1):
        cell = ws.cell(row=3, column=c_idx, value=str(col).replace("_"," ").title())
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[3].height = 22

    # Data rows
    for r_idx, row in enumerate(df_data.itertuples(index=False), 4):
        fill = CELL_FILL_ODD if r_idx % 2 == 0 else CELL_FILL_EVEN
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = CELL_FONT
            cell.fill      = fill
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal="right" if isinstance(val,(int,float)) else "left",
                                       vertical="center")
            # Number formatting
            if isinstance(val, float):
                cell.number_format = "#,##0.00"
            elif isinstance(val, int):
                cell.number_format = "#,##0"

    # Auto-fit columns
    for c_idx, col in enumerate(df_data.columns, 1):
        max_len = max(
            len(str(col)) + 4,
            df_data[col].astype(str).str.len().max() if len(df_data) else 10
        )
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 35)

    # Freeze panes
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False

# ── Cover / Index Sheet ───────────────────────────────────
ws_cover = wb.create_sheet(title="📊 Index", index=0)
ws_cover.sheet_properties.tabColor = "ffd166"
ws_cover.sheet_view.showGridLines = False

cover_data = [
    ("", ""),
    ("", "E-COMMERCE BUSINESS INTELLIGENCE PLATFORM"),
    ("", "SQL Analytics Report"),
    ("", f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}"),
    ("", ""),
    ("", "QUERY INDEX"),
    ("", "─" * 60),
]
for qnum, (key, (title, _)) in enumerate(results.items(), 1):
    cover_data.append((f"Q{qnum:02d}", f"{title}"))

for r, (label, text) in enumerate(cover_data, 1):
    ws_cover.cell(row=r, column=2, value=label).font = Font(color="ffd166", bold=True, name="Calibri", size=11)
    cell = ws_cover.cell(row=r, column=3, value=text)
    if r == 2:
        cell.font = Font(color="00c8ff", bold=True, name="Calibri", size=18)
    elif r == 3:
        cell.font = Font(color="e8e8f0", name="Calibri", size=13)
    elif r == 4:
        cell.font = Font(color="aaaacc", name="Calibri", size=10)
    elif r == 6:
        cell.font = Font(color="ffd166", bold=True, name="Calibri", size=12)
    else:
        cell.font = Font(color="e8e8f0", name="Calibri", size=10)
    ws_cover.cell(row=r, column=2).fill = hex_fill("0d0d1a")
    ws_cover.cell(row=r, column=3).fill = hex_fill("0d0d1a")

for row in ws_cover.iter_rows():
    for cell in row:
        cell.fill = hex_fill("0d0d1a")

ws_cover.column_dimensions["A"].width = 3
ws_cover.column_dimensions["B"].width = 8
ws_cover.column_dimensions["C"].width = 70

# ── Write each query to its own sheet ─────────────────────
for key, (title, df_res) in results.items():
    short = key[:31]
    write_query_sheet(wb, short, title, df_res)

# ── Executive Summary Sheet (special formatting) ───────────
exec_title, exec_df = results["Q25_Executive_Summary"]
ws_exec = wb["Q25_Executive_Summary"]
ws_exec.sheet_properties.tabColor = "ff4d6d"

wb.save(XLS_OUT)
print(f"  Excel Report → {XLS_OUT}")

# ═══════════════════════════════════════════════════════════
#  STEP 4 — SQL Visualisation Summary (PNG)
# ═══════════════════════════════════════════════════════════
print("\n  [4/5] Generating SQL analytics visualisation …")

_, df_monthly  = results["Q04_Monthly_Revenue_Trend"]
_, df_category = results["Q08_Revenue_by_Category"]
_, df_state    = results["Q13_Top_States_Revenue"]
_, df_status   = results["Q02_Order_Status_Analysis"]
_, df_channel  = results["Q16_Sales_Channel_Revenue"]
_, df_fulfilm  = results["Q17_Fulfilment_Analysis"]
_, df_return   = results["Q23_Return_Rate_by_Category"]
_, df_exec     = results["Q25_Executive_Summary"]

fig = plt.figure(figsize=(24, 20))
fig.patch.set_facecolor(BG)
gs  = gridspec.GridSpec(4, 3, figure=fig, hspace=0.50, wspace=0.38,
                        top=0.94, bottom=0.04, left=0.05, right=0.97)

plt.suptitle("SQL Analytics Summary — E-Commerce BI Platform",
             color="white", fontsize=18, fontweight="bold", y=0.97)

# ── Monthly Revenue ───────────────────────────────────────
ax1 = fig.add_subplot(gs[0, :])
ax1.set_facecolor(PANEL)
x = range(len(df_monthly))
ax1.bar(x, df_monthly["total_revenue"] / 1e6, color="#00c8ff", alpha=0.65, label="Revenue (₹M)")
ax2 = ax1.twinx()
ax2.plot(x, df_monthly["total_orders"], color="#ff4d6d", lw=2.5, marker="o", ms=6, label="Orders")
ax1.set_xticks(list(x))
ax1.set_xticklabels(
    [f"{row['month_name']} {row['year']}" for _, row in df_monthly.iterrows()],
    color="white", rotation=30, fontsize=9)
ax1.set_ylabel("Revenue (₹ Millions)", color="#00c8ff")
ax2.set_ylabel("Orders", color="#ff4d6d")
ax1.set_title("Monthly Revenue & Order Trend (SQL Query Q04)", color="white", fontsize=13, fontweight="bold")
ax1.tick_params(colors="white", axis="y")
ax2.tick_params(colors="#ff4d6d", axis="y")
ax1.set_facecolor(PANEL)
# Growth annotations
if "mom_growth_pct" in df_monthly.columns:
    for i, row in df_monthly.iterrows():
        if pd.notna(row.get("mom_growth_pct",None)) and row["mom_growth_pct"] != 0:
            color = "#06d6a0" if row["mom_growth_pct"] > 0 else "#ff4d6d"
            ax2.annotate(f"{row['mom_growth_pct']:+.1f}%",
                         xy=(i, df_monthly["total_orders"].iloc[i]),
                         xytext=(0, 10), textcoords="offset points",
                         ha="center", color=color, fontsize=7, fontweight="bold")
lines1, labels1 = ax1.get_legend_handles_labels()
lines2, labels2 = ax2.get_legend_handles_labels()
ax1.legend(lines1+lines2, labels1+labels2, labelcolor="white", facecolor="#1c1c3a", edgecolor="#333", fontsize=9)
for sp in ax1.spines.values(): sp.set_edgecolor("#333")

# ── Revenue by Category ───────────────────────────────────
ax3 = fig.add_subplot(gs[1, 0])
ax3.set_facecolor(PANEL)
bars = ax3.barh(df_category["category"][::-1], df_category["total_revenue"][::-1] / 1e6,
                color=PAL[:len(df_category)])
ax3.set_xlabel("Revenue (₹ Millions)", color="white")
ax3.set_title("Revenue by Category\n(Q08)", color="white", fontsize=11, fontweight="bold")
ax3.tick_params(colors="white")
for sp in ax3.spines.values(): sp.set_edgecolor("#333")
for bar, pct in zip(bars, df_category["revenue_pct"][::-1]):
    ax3.text(bar.get_width() + 0.05, bar.get_y() + bar.get_height()/2,
             f"{pct:.1f}%", va="center", color="white", fontsize=8)

# ── Top States ────────────────────────────────────────────
ax4 = fig.add_subplot(gs[1, 1])
ax4.set_facecolor(PANEL)
top10 = df_state.head(10)
colors_g = plt.cm.cool(np.linspace(0.2, 0.9, len(top10)))
ax4.barh(top10["ship_state"][::-1], top10["total_revenue"][::-1] / 1e6, color=colors_g)
ax4.set_xlabel("Revenue (₹ Millions)", color="white")
ax4.set_title("Top 10 States by Revenue\n(Q13)", color="white", fontsize=11, fontweight="bold")
ax4.tick_params(colors="white")
for sp in ax4.spines.values(): sp.set_edgecolor("#333")

# ── Order Status Donut ────────────────────────────────────
ax5 = fig.add_subplot(gs[1, 2])
ax5.set_facecolor(PANEL)
wedges, texts, auto = ax5.pie(
    df_status["order_count"], labels=None, autopct="%1.1f%%",
    colors=PAL[:len(df_status)], startangle=90,
    pctdistance=0.80, wedgeprops=dict(width=0.55, edgecolor=BG, linewidth=2))
for t in auto: t.set_color("white"); t.set_fontsize(8)
ax5.legend(df_status["status"], loc="lower center", fontsize=7,
           labelcolor="white", facecolor="#1c1c3a", edgecolor="#333",
           bbox_to_anchor=(0.5, -0.18), ncol=2)
ax5.set_title("Order Status Distribution\n(Q02)", color="white", fontsize=11, fontweight="bold")

# ── Return Rate by Category ───────────────────────────────
ax6 = fig.add_subplot(gs[2, 0])
ax6.set_facecolor(PANEL)
ax6.barh(df_return["category"][::-1], df_return["return_rate_pct"][::-1],
         color="#ff4d6d")
ax6.set_xlabel("Return Rate (%)", color="white")
ax6.set_title("Return Rate by Category\n(Q23)", color="white", fontsize=11, fontweight="bold")
ax6.tick_params(colors="white")
ax6.axvline(df_return["return_rate_pct"].mean(), color="#ffd166", linestyle="--", lw=1.5,
            label=f"Avg {df_return['return_rate_pct'].mean():.1f}%")
ax6.legend(labelcolor="white", facecolor="#1c1c3a", edgecolor="#333", fontsize=8)
for sp in ax6.spines.values(): sp.set_edgecolor("#333")

# ── Channel Revenue ───────────────────────────────────────
ax7 = fig.add_subplot(gs[2, 1])
ax7.set_facecolor(PANEL)
ax7.bar(df_channel["sales_channel"], df_channel["total_revenue"] / 1e6,
        color=["#00c8ff","#c77dff"])
ax7.set_ylabel("Revenue (₹ Millions)", color="white")
ax7.set_title("Sales Channel Revenue\n(Q16)", color="white", fontsize=11, fontweight="bold")
ax7.tick_params(colors="white")
for sp in ax7.spines.values(): sp.set_edgecolor("#333")

# ── Fulfilment Analysis ───────────────────────────────────
ax8 = fig.add_subplot(gs[2, 2])
ax8.set_facecolor(PANEL)
x8 = range(len(df_fulfilm))
w  = 0.35
ax8.bar([i - w/2 for i in x8], df_fulfilm["total_revenue"] / 1e6,
        width=w, color="#06d6a0", label="Revenue (₹M)")
ax8b = ax8.twinx()
ax8b.bar([i + w/2 for i in x8], df_fulfilm["cancel_rate_pct"],
         width=w, color="#ff4d6d", label="Cancel Rate %")
ax8.set_xticks(list(x8))
ax8.set_xticklabels(df_fulfilm["fulfilment"], color="white")
ax8.set_ylabel("Revenue (₹M)", color="#06d6a0")
ax8b.set_ylabel("Cancel Rate %", color="#ff4d6d")
ax8.set_title("Fulfilment Analysis\n(Q17)", color="white", fontsize=11, fontweight="bold")
ax8.tick_params(colors="white", axis="y")
ax8b.tick_params(colors="#ff4d6d", axis="y")
for sp in ax8.spines.values(): sp.set_edgecolor("#333")

# ── Executive Summary Table ───────────────────────────────
ax9 = fig.add_subplot(gs[3, :])
ax9.set_facecolor(PANEL)
ax9.axis("off")
ax9.set_title("Executive KPI Summary (Q25)", color="white", fontsize=12, fontweight="bold", pad=10)

kpi_data = df_exec.values.tolist()
n_cols   = 4
n_rows   = int(np.ceil(len(kpi_data) / n_cols))
col_w    = 1.0 / n_cols

for idx, (metric, value) in enumerate(kpi_data):
    col = idx % n_cols
    row = idx // n_cols
    x   = col * col_w + 0.01
    y   = 1 - (row + 1) * (1.0 / (n_rows + 1))
    box = mpatches.FancyBboxPatch((x, y - 0.06), col_w - 0.02, 0.10,
                                  boxstyle="round,pad=0.01",
                                  facecolor="#1c1c3a", edgecolor="#00c8ff",
                                  linewidth=1.5, transform=ax9.transAxes, clip_on=True)
    ax9.add_patch(box)
    ax9.text(x + (col_w - 0.02)/2, y + 0.02,
             f"{value:,}" if isinstance(value,(int,float)) else str(value),
             ha="center", va="center", color="#00c8ff",
             fontsize=13, fontweight="bold", transform=ax9.transAxes)
    ax9.text(x + (col_w - 0.02)/2, y - 0.04,
             str(metric), ha="center", va="center", color="#aaaacc",
             fontsize=8, transform=ax9.transAxes)

viz_path = os.path.join(RPT_DIR, "sql_analytics_summary.png")
plt.savefig(viz_path, dpi=150, bbox_inches="tight", facecolor=BG)
plt.close()
print(f"  SQL Analytics PNG → {viz_path}")

# ═══════════════════════════════════════════════════════════
#  STEP 5 — Write SQL File Summary
# ═══════════════════════════════════════════════════════════
print("\n  [5/5] Writing query catalogue …")

catalogue_path = os.path.join(SQL_DIR, "query_catalogue.md")
with open(catalogue_path, "w") as f:
    f.write("# SQL Query Catalogue\n")
    f.write("## E-Commerce BI Platform — 25 Business Analysis Queries\n\n")
    f.write(f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*\n\n")
    f.write("| # | Query Key | Business Title | Rows Returned |\n")
    f.write("|---|-----------|----------------|---------------|\n")
    for qnum, (key, (title, df_res)) in enumerate(results.items(), 1):
        f.write(f"| {qnum:02d} | `{key}` | {title} | {len(df_res)} |\n")
    f.write("\n---\n\n## Key Business Insights\n\n")

    # Auto-generate insights from results
    _, df_kpi = results["Q01_Overall_KPIs"]
    _, df_cat = results["Q08_Revenue_by_Category"]
    _, df_st  = results["Q13_Top_States_Revenue"]
    _, df_ret = results["Q23_Return_Rate_by_Category"]

    f.write(f"- **Total Revenue**: ₹{df_kpi['total_revenue_inr'].iloc[0]:,.2f}\n")
    f.write(f"- **Total Orders**: {int(df_kpi['total_orders'].iloc[0]):,}\n")
    f.write(f"- **Avg Order Value**: ₹{df_kpi['avg_order_value'].iloc[0]:,.2f}\n")
    f.write(f"- **Top Category**: {df_cat['category'].iloc[0]} "
            f"(₹{df_cat['total_revenue'].iloc[0]:,.0f})\n")
    f.write(f"- **Top State**: {df_st['ship_state'].iloc[0]} "
            f"(₹{df_st['total_revenue'].iloc[0]:,.0f})\n")
    f.write(f"- **Highest Return Rate**: {df_ret['category'].iloc[0]} "
            f"({df_ret['return_rate_pct'].iloc[0]}%)\n")

print(f"  Query Catalogue → {catalogue_path}")

print("\n" + "="*65)
print("  ✅  SQL ANALYTICS RUNNER COMPLETE")
print(f"  Queries Executed : 25")
print(f"  Excel Report     : {XLS_OUT}")
print(f"  Visualisation    : {viz_path}")
print(f"  SQLite DB        : {DB_PATH}")
print("="*65 + "\n")
